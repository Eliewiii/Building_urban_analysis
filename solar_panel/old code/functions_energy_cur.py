# -*- coding: utf-8 -*-
"""
Created on Sun Apr 23 11:14:07 2023

@author: Julius Jandl
"""


import json
import numpy as np
import sys
import math
import openpyxl

from functions_dmfa_cur import *
np.set_printoptions(threshold=sys.maxsize)

datafile = openpyxl.load_workbook(filename = "dmfa model 2.xlsx", data_only=True)
energyfile = openpyxl.load_workbook(filename = "energydata.xlsx")
ds_inputs = datafile["assumptions"]
area_file = datafile["area_input_2023_01_18"]
efficiency_file = datafile["Efficiency development"]

nb_of_years = 100
list_b_type = [area_file.cell(i,3).value for i in range(4,5)]
list_tech = ["rooftech", "envtech"]
list_scenario = ["roof_scenario", "env_scenario"]
list_wb = ["regular loss", "early loss"]
list_level = ["efficient", "aesthetic"]

lt_PV = ds_inputs.cell(28,2).value # average lifetime of PV Panels
first_year = ds_inputs.cell(23,2).value
list_years = [i for i in range(first_year,first_year + nb_of_years)]
list_life = [i for i in range(0,nb_of_years)]
wb_shape_EL = ds_inputs.cell(26,2).value # EarLy Loss scenario
wb_shape_RL = ds_inputs.cell(27,2).value # Regular Loss scenario

# area_factors = {"roof" : 0.32, "envelop_600" : 0.8, "envelop_800" : 0.8}
area_factor_roof = 0.32 # paper Vardimon
area_factor_env = 0.7/0.8 # 0.8 is in excel data
pv_efficiency = 0.2 # already in excel data

# PV parameter
# rooftech = Mitrex roof panels
# envtech eff = metsolar glass black 
# envtech aesth = metsolar glass beige

# lists of increasing efficiency over time for new panels
list_eff_rooftech = [efficiency_file.cell(i,7).value for i in range(2,102)]
list_eff_envtech_eff = [efficiency_file.cell(i,8).value for i in range(2,102)]
list_eff_envtech_aesth = [efficiency_file.cell(i,9).value for i in range(2,102)]

mass_conv = 0.01575 # area to mass conversion, value in ton/m^2
area_to_module = 1.46 # area per module in m^2
weight_to_module = 0.023 # weight per module in tons
pv_performance_ratio = 0.75 # coefficient for losses from cables, conversion, ...

energy_efficiency = {"efficient": {"rooftech" : list_eff_rooftech, "envtech" : list_eff_envtech_eff},
                     "aesthetic": {"rooftech" : list_eff_rooftech, "envtech" : list_eff_envtech_aesth}} # PV conversion efficiency

# construction process, phased construction for neighbourhoods in smaller portions of buildings
dic_construction_simple = {"dtrain" : {"total_nb_of_buildings" : 4, "nb_buildings_parallel" : 4, "construction_time" : 1},
                "h" : {"total_nb_of_buildings" : 17, "nb_buildings_parallel" : 17, "construction_time" : 1},
                "L" : {"total_nb_of_buildings" : 20, "nb_buildings_parallel" : 20, "construction_time" : 1},
                "R" : {"total_nb_of_buildings" : 21, "nb_buildings_parallel" : 21, "construction_time" : 1},
                "t" : {"total_nb_of_buildings" : 19, "nb_buildings_parallel" : 19, "construction_time" : 1}}

dic_construction_phased = {"dtrain" : {"total_nb_of_buildings" : 4, "nb_buildings_parallel" : 1, "construction_time" : 1},
                "h" : {"total_nb_of_buildings" : 17, "nb_buildings_parallel" : 5, "construction_time" : 1},
                "L" : {"total_nb_of_buildings" : 20, "nb_buildings_parallel" : 4, "construction_time" : 1},
                "R" : {"total_nb_of_buildings" : 21, "nb_buildings_parallel" : 4, "construction_time" : 1},
                "t" : {"total_nb_of_buildings" : 19, "nb_buildings_parallel" : 4, "construction_time" : 1}}

def calc_wb(nb_of_years, lifetime, shape):
    """ Calculate Weibull distribution values """
    list_wb = []
    for i in range(0, nb_of_years):
        val = shape*i**(shape-1)/(lifetime**shape)*math.exp(-(i/lifetime)**shape)
        list_wb.append(val)

    return(list_wb)

def calc_wb_cumu(nb_of_years, lifetime, shape):
    """ calculate cumulative Weibull distribution values """
    list_wb_cumu = []
    for i in range(0, nb_of_years):
        val = (1-np.exp(-(i / lifetime)**shape))
        list_wb_cumu.append(val)
        
    return(list_wb_cumu)

def initialize_dic_energy(list_level, list_tech, list_b_type):
    """ Make a dictionary to put in energy production values """
    
    # intialize area dictionary
    dic_energy = {}
    for year in range(nb_of_years):
        dic_energy[year] = {}
        for b_type in list_b_type: 
            dic_energy[year][b_type] = {"roof_scenario": {}, "env_scenario": {}, "total_scenario" : {}}
            
            # Roof
            dic_energy[year][b_type]["roof_scenario"] = {}
            dic_energy[year][b_type]["roof_scenario"][list_tech[0]] = {"energy_production": 0.}
                
            # envelope
            dic_energy[year][b_type]["env_scenario"] = {}
            dic_energy[year][b_type]["env_scenario"][list_tech[1]] = {"energy_production": 0.}
                    
            ## Total
            dic_energy[year][b_type]["total_scenario"] = {}  # Initialization
            for tech in list_tech:
                dic_energy[year][b_type]["total_scenario"][tech] = {"energy_production": 0.}
    
    return dic_energy

def initialize_dic_area(list_b_type):
    """ Make a dictionary to put in area data for stock calculation """
    
    # intialize area dictionary
    dic_area = {}
    for year in range(nb_of_years):
        dic_area[year] = {}
        for b_typ in list_b_type: 
            dic_area[year][b_typ] = {"roof_scenario": {}, "env_scenario": {}}
            
            # Roof
            dic_area[year][b_typ]["roof_scenario"] = {"area": 0.}
                
            # envelope
            dic_area[year][b_typ]["env_scenario"] = {"area" : 0.}
    
    return dic_area

def initialize_dic_irradiation(list_b_type):
    """ make a dictionary to put in irradiation values per scenario """
    
    dic_irradiation = {}
    for year in range(nb_of_years):
        dic_irradiation[year] = {}
        for b_type in list_b_type: 
            dic_irradiation[year][b_type] = {"roof_scenario": {}, "env_scenario": {}}
            
            # Roof
            dic_irradiation[year][b_type]["roof_scenario"] = {"irradiation": 0.}
                
            # envelope
            dic_irradiation[year][b_type]["env_scenario"] = {"irradiation" : 0.}
        
    return dic_irradiation

def initialize_dic_age(list_b_type):
    """ make a dictionary to put in age values per scenario, necessary only if class is used """
    
    dic_age = {}
    for year in range(0, nb_of_years):
        dic_age[year] = {}
        for b_type in list_b_type:
            dic_age[year][b_type] = {}
            for tech in list_tech:
                dic_age[year][b_type][tech] = {"installed modules": 0.}
    
    return dic_age
              
def calc_area_factor(dic_construction, construction_scenario, b_type):
    """ reducing area in early years of less existing construction """
    
    buildtype = b_type.split("_")[0] # only get name of b_type without angle of orientation
    if construction_scenario == "simple":
        dic_construction = dic_construction_simple
    elif construction_scenario == "phased":
        dic_construction = dic_construction_phased

    #parameters for construction processs
    total_nb_of_buildings = dic_construction[buildtype]["total_nb_of_buildings"]
    nb_buildings_parallel = dic_construction[buildtype]["nb_buildings_parallel"]
    construction_time = dic_construction[buildtype]["construction_time"]
    length = (total_nb_of_buildings-nb_buildings_parallel)/nb_buildings_parallel*construction_time

    if total_nb_of_buildings % nb_buildings_parallel != 0:
        phases = [i for i in range(0, round(length), construction_time)]
    elif total_nb_of_buildings % nb_buildings_parallel == 0:
        phases = [i for i in range(0, int(length) + 1, construction_time)]

    buildings = []
    nb_built = nb_buildings_parallel
    for year in range(1, nb_of_years + 1):
        if year in phases:
            if nb_built < total_nb_of_buildings <= nb_built + nb_buildings_parallel:
                nb_built = total_nb_of_buildings
            elif nb_built + nb_buildings_parallel < total_nb_of_buildings:
                nb_built = nb_built + nb_buildings_parallel
     
        elif year > phases[-1] + 1:
            nb_built = total_nb_of_buildings
            
        buildings.append(nb_built)
     
    buildings.insert(0, nb_buildings_parallel)
    buildings = buildings[0:100]
    area_factor = [i/total_nb_of_buildings for i in buildings]

    return area_factor
              
def calc_area(dic_area, construction_scenario, list_scenario, list_b_type, datasheet, rows, columns):
    """ creates array with all PV-area inputs for each scenario and writes the values in the dic_area """
    
    if construction_scenario == "simple":
        dic_construction = dic_construction_simple
    elif construction_scenario == "phased":
        dic_construction = dic_construction_phased

    arr_area = np.zeros((len(list_b_type), len(list_scenario)))
    for row in range(rows[0], rows[1]):
        for column in range(columns[0], columns[1]):
            if column == columns[0]:
                arr_area[row-rows[0],column-columns[0]] = datasheet.cell(row,column).value * area_factor_roof
            else:
                arr_area[row-rows[0],column-columns[0]] = datasheet.cell(row,column).value * area_factor_env
    r = 0
    c = 0
    for scenario in list_scenario:
        for b_type in list_b_type:
            # print(b_type)
            # print(scenario)
            factorlist = calc_area_factor(dic_construction, construction_scenario, b_type)
            # print(factorlist)
            for year in range(nb_of_years):
                if scenario == "roof_scenario":
                    dic_area[year][b_type]["roof_scenario"]["area"] = arr_area[r,c] * factorlist[year]
                else: 
                    dic_area[year][b_type]["env_scenario"]["area"] = arr_area[r, c] * factorlist[year]
            
            r = r + 1
        c = c + 1
        r = 0
    return arr_area

def calc_energy_arr(list_b_type, construction_scenario, dic_irradiation, datasheet, rows, columns, list_scenario):
    
    if construction_scenario == "simple":
        dic_construction = dic_construction_simple
    elif construction_scenario == "phased":
        dic_construction = dic_construction_phased
    
    arr_irradiation = np.zeros((len(list_b_type), len(list_scenario)))
    for row in range(rows[0], rows[1]):
        for column in range(columns[0], columns[1]):
            if column == columns[0]:
                arr_irradiation[row-rows[0],column-columns[0]] = datasheet.cell(row,column).value*area_factor_roof
            else:
                arr_irradiation[row-rows[0],column-columns[0]] = datasheet.cell(row,column).value*area_factor_env
    
    r = 0
    c = 0
    for scenario in list_scenario:
        for b_type in list_b_type:
            factorlist = calc_area_factor(dic_construction, construction_scenario, b_type)
            
            for year in range(nb_of_years):
                if scenario == "roof_scenario":
                    dic_irradiation[year][b_type]["roof_scenario"]["irradiation"] = arr_irradiation[r, c] * factorlist[year]
                elif scenario == "env_scenario": 
                    dic_irradiation[year][b_type]["env_scenario"]["irradiation"] = arr_irradiation[r, c] * factorlist[year]
                else: 
                    print("error")
                    
            r = r + 1
        c = c + 1
        r = 0
    return arr_irradiation, dic_irradiation

def calc_degradation(age):
    """calculates degradation for modules according to their age"""
    if age == 0:
        deg = 0
    elif age == 1:
        deg = 0.02 
    elif age > 1:
        deg = 0.02 + 0.005*(age-1)
        
    return deg

def calc_energy_production(dic_result_per_year, dic_area, construction_scenario, list_level, list_b_type, datasheet, rows, columns, wb, level):
    """ calculate energy output of PV modules in GWh/year """
    
    dic_irradiation = initialize_dic_irradiation(list_b_type)
    dic_energy = initialize_dic_energy(list_level, list_tech, list_b_type)
    arr_irradiation, dic_irradiation = calc_energy_arr(list_b_type, construction_scenario, dic_irradiation, datasheet, rows, columns, list_scenario)
    for b_type in list(dic_result_per_year[0].keys()):
        list_sum = []
        for var in list_scenario: 
            if var == "roof_scenario":
                tech = "rooftech"
            elif var == "env_scenario":
                tech = "envtech"
                
            waste_arr, waste_modules_arr, list_inflow, list_inflow_modules = calc_cohorts(dic_result_per_year, b_type, var, tech, wb, level, nb_of_years)
            for year in range(nb_of_years):
                if var == "roof_scenario" or var == "env_scenario":
                    list_energy = []
                    module_check = []
                    left_weight_list, left_module_list, waste_list, waste_module_list = calc_left_modules(dic_result_per_year, year, b_type, var, tech, wb, level, nb_of_years)

                    for i in range(0,year + 1):
                       modules_total_on_building = dic_result_per_year[year][b_type][var][tech]["modules"]
                       modules_total_in_year_i = dic_result_per_year[i][b_type][var][tech]["modules"]

                       #from initial year
                       if i == 0:
                           failed_modules_from_init = 0
                           modules_init = dic_result_per_year[0][b_type][var][tech]["modules"]
                           for x in range(0,year):
                               failed_modules_from_init += waste_modules_arr[0,x]
                               
                           age = year    
                           deg = calc_degradation(age)
                           modules_from_i_in_year = modules_init - failed_modules_from_init
                           if modules_from_i_in_year <= 0:
                               modules_from_i_in_year = 0
                         
                       elif i > 0:
                            
                           #modules from waste
                           new_modules_in_year_i = np.sum(waste_modules_arr, axis = 0)[i]
                           age = year - i
                           deg = calc_degradation(age)
                        
                           #modules from new buildings
                           new_modules_from_construction = 0
                           if dic_result_per_year[i][b_type][var][tech]["modules"] > dic_result_per_year[i - 1][b_type][var][tech]["modules"]:
                               new_modules_from_construction = dic_result_per_year[i][b_type][var][tech]["modules"] - dic_result_per_year[i - 1][b_type][var][tech]["modules"]

                           modules_from_i_in_year = new_modules_in_year_i + new_modules_from_construction - waste_module_list[i]
                           if modules_from_i_in_year < 0:
                               modules_from_i_in_year = 0
                           
                       energy = (modules_from_i_in_year/modules_total_in_year_i * dic_irradiation[i][b_type][var]["irradiation"]
                                  / pv_efficiency * energy_efficiency[level][tech][i] * pv_performance_ratio / 1000 * (1-deg))
                       list_energy.append(energy)

                if var == "total_scenario":
                    pass
                else:
                    # print("energy", year, b_type, var, sum(list_energy))    
                    dic_energy[year][b_type][var][tech]["energy_production"] = sum(list_energy)
                
                list_sum.append(sum(list_energy))

    return dic_energy, dic_irradiation

def compute_total_scenarios_energy(dic_energy):
    """ Compute the stock and the waste of combined roof and envelop scenarios (and eventually others if needed """
    #  loop over all he years ( we add list before to convert it as it's not exactly a list and we cannot do a loop on it
    for year in list(dic_energy.keys()):
            for b_type in list(dic_energy[year].keys()):
                for tech in list(dic_energy[year][b_type]["total_scenario"].keys()):
                    # Roof
                    try:
                        dic_energy[year][b_type]["roof_scenario"][tech]["energy_production"]
                        # print(dic_energy[year][b_type]["roof_scenario"][tech]["energy_production"])
                    except:
                        roof_energy_production = 0
                    else:
                        roof_energy_production = dic_energy[year][b_type]["roof_scenario"][tech]["energy_production"]

                    # Envelop
                    try:
                        dic_energy[year][b_type]["env_scenario"][tech]["energy_production"]
                    except:
                        env_energy_production = 0
                    else:
                        env_energy_production = dic_energy[year][b_type]["env_scenario"][tech]["energy_production"]
                        
                    # Total
                    dic_energy[year][b_type]["total_scenario"][tech]["energy_production"] = roof_energy_production + env_energy_production

    return dic_energy

def plot_energy(dic_energy, b_type, construction_scenario):
    """plots energy production for each year"""
    energy = []
    for i in range(nb_of_years):
        energy_production_roof = dic_energy[i][b_type]["roof_scenario"]["rooftech"]["energy_production"]
        energy_production_env = dic_energy[i][b_type]["env_scenario"]["envtech"]["energy_production"]
        energy_production = energy_production_roof + energy_production_env
        energy.append(energy_production)
            
    fig, ax = plt.subplots(figsize=(9, 6))
    if construction_scenario == "simple":
        ax.plot(list_life[:], energy, color = "g")
    elif construction_scenario == "phased":
        ax.plot(list_life[:], energy, color = "r")
    ax.set_xlabel('Years')
    ax.set_ylabel('Energy generation', color='k')
    ax.set_ylim(0, 15)
    ax.set_xlim(-0.5,100)
    plt.title("Total Energy Production " + b_type)
    plt.rcParams.update({'font.size': 14})
    plt.legend(loc="upper left")
    plt.show    

def write_energy_data(dic_energy, wb, level, year):
    """writes energy data to .xlsx file"""
    
    ws = energyfile.create_sheet("energy " + wb + " " + level)
    ws.cell(1,1).value = "Energy in GWh/year"
    n = 2   
    m = 2
    
    for b_type in list_b_type:
        for tech in list_tech:
            for y in range(year):
                res = dic_energy[y][b_type]["total_scenario"][tech]["energy_production"]
                cell_name = ws.cell(n, 1)   
                if tech == "rooftech":
                    cell_energy = ws.cell(n, m)
                    cell_label = ws.cell(1,m)
    
                elif tech == "envtech":
                    cell_energy = ws.cell(n, m+1)
                    cell_label = ws.cell(1,m+1)
                    
                cell_name.value = y
                cell_energy.value = res
                cell_label.value = tech
    
                n += 1
        m += 2
        
    energyfile.save("energydata.xlsx")
    
def run_energy(dic_energy, dic_result_per_year, wb, level, construction):
    """runs all calculations and writes data to .json files"""
    
    dic_area = initialize_dic_area(list_b_type)
    dic_energy = initialize_dic_energy(list_level, list_tech, list_b_type)
    arr_area = calc_area(dic_area, construction, list_scenario, list_b_type, area_file, [4,5], [4,6])
    dic_energy, dic_irradiation = calc_energy_production(dic_result_per_year, dic_area, construction, list_level, list_b_type, area_file, [4,5], [7,9], wb, level)

    path = "irradiation.json"
    with open(path, "w") as out_file:
        json.dump(dic_irradiation, out_file, indent = 4)
        
    path = "energy_single.json"
    with open(path, "w") as out_file:
        json.dump(dic_energy, out_file, indent = 4)

    compute_total_scenarios_energy(dic_energy)
        
    path = "energy_complete.json"
    with open(path, "w") as out_file:
        json.dump(dic_energy, out_file, indent = 4)
    
    for b_type in list_b_type:
        plot_energy(dic_energy, b_type, construction)

    return dic_energy

wb_cumu_RL = calc_wb_cumu(nb_of_years, lt_PV, wb_shape_RL)
wb_cumu_EL = calc_wb_cumu(nb_of_years, lt_PV, wb_shape_EL)
dic_wb = {"regular loss" : wb_cumu_RL, "early loss" : wb_cumu_EL}