# -*- coding: utf-8 -*-
"""
Created on Mon Jan 16 15:55:18 2023

@author: Julius Jandl

This version uses the cumulative Weibull function to calculate waste streams.
"""

import json
import numpy as np
import math
import openpyxl
from openpyxl import Workbook
import matplotlib.pyplot as plt
import pandas as pd
import copy
import random
import sys
np.set_printoptions(threshold=sys.maxsize)

datafile = openpyxl.load_workbook(filename = "dmfa model 2.xlsx", data_only=True)
ds_inputs = datafile["assumptions"]
area_file = datafile["area_input_2023_01_18"]
efficiency_file = datafile["Efficiency development"]


nb_of_years = 100
list_b_type = [area_file.cell(i, 3).value for i in range(4, 5)]
list_scenario = ["roof_scenario", "env_scenario"]
list_tech = ["rooftech", "envtech"]
list_wb = ["regular loss", "early loss"]
list_level = ["efficient", "aesthetic"]

wb_shape_EL = ds_inputs.cell(26,2).value # EarLy Loss scenario
wb_shape_RL = ds_inputs.cell(27,2).value # Regular Loss scenario
lt_PV = ds_inputs.cell(28,2).value # average lifetime of PV Panels
first_year = ds_inputs.cell(23,2).value
list_years = [i for i in range(first_year,first_year + nb_of_years)]
list_life = [i for i in range(0,nb_of_years)]

# area_factors = {"roof" : 0.32, "envelop_600" : 0.8, "envelop_800" : 0.8}
area_factor_roof = 0.32 # paper Vardimon
area_factor_env = 0.7/0.8 # 0.8 is in excel data
pv_efficiency = 0.2 # already in excel data

# PV parameter
#rooftech = Mitrex roof panels
# envtech eff = metsolar glass black 
# envtech aesth = metsolar glass beige

list_eff_rooftech = [efficiency_file.cell(i,7).value for i in range(2,102)]
list_eff_envtech_eff = [efficiency_file.cell(i,8).value for i in range(2,102)]
list_eff_envtech_aesth = [efficiency_file.cell(i,9).value for i in range(2,102)]

mass_conv = 0.01575 # area to mass conversion, value in ton/m^2
area_to_module = 1.46 # area per module in m^2
weight_to_module = 0.023 # weight per module in tons
pv_performance_ratio = 0.75 # coefficient for losses from cables, conversion, ...

energy_efficiency = {"efficient": {"rooftech" : list_eff_rooftech, "envtech" : list_eff_envtech_eff},
                     "aesthetic": {"rooftech" : list_eff_rooftech, "envtech" : list_eff_envtech_aesth}} # PV conversion efficiency


# construction process, phased construction for neighbourhoods in smaller portions of buildings
dic_construction_simple = {"dtrain" : {"total_nb_of_buildings" : 4, "nb_buildings_parallel" : 4, "construction_time" : 1},
                "h" : {"total_nb_of_buildings" : 17, "nb_buildings_parallel" : 17, "construction_time" : 1},
                "L" : {"total_nb_of_buildings" : 20, "nb_buildings_parallel" : 20, "construction_time" : 1},
                "R" : {"total_nb_of_buildings" : 21, "nb_buildings_parallel" : 21, "construction_time" : 1},
                "t" : {"total_nb_of_buildings" : 19, "nb_buildings_parallel" : 19, "construction_time" : 1}}

dic_construction_phased = {"dtrain" : {"total_nb_of_buildings" : 4, "nb_buildings_parallel" : 1, "construction_time" : 1},
                "h" : {"total_nb_of_buildings" : 17, "nb_buildings_parallel" : 5, "construction_time" : 1},
                "L" : {"total_nb_of_buildings" : 20, "nb_buildings_parallel" : 4, "construction_time" : 1},
                "R" : {"total_nb_of_buildings" : 21, "nb_buildings_parallel" : 4, "construction_time" : 1},
                "t" : {"total_nb_of_buildings" : 19, "nb_buildings_parallel" : 4, "construction_time" : 1}}

class Panel:
    
    def __init__(self,index):
        self.index = index
        self.failure_scenario_list = []
        
    def compute_failure_scenario(self,lambda_para,k,nb_of_year_building,number_of_scenario_to_compute):
        """
        Computes different failure scenarios for the panel and creates a list for each one
        """
        
        total_list = []
        for i in range(0,number_of_scenario_to_compute):
            panel_list = []
            while True:
                y = random.random()
                time_of_failure = math.ceil(lambda_para * (-(1/k) * math.log(1 - y))**(1/k))
                if sum(panel_list) + time_of_failure <= nb_of_year_building:
                    panel_list.append(time_of_failure)
                else:
                    time_of_failure = nb_of_year_building - sum(panel_list) 
                    panel_list.append(time_of_failure)
                    break
            total_list.append(panel_list)

def make_total_scenario(list_scenario_roof, list_scenario_env):
    """ Make total scenarios with total roof and envelop combinations """
    # Initialize the list that will host the new total scenarios
    list_total_scenario = []
    # Creation of the new scenario
    for scenario_roof in list_scenario_roof:  # roof scenario
        for scenario_env in list_scenario_env:  # envelop scenario
            list_total_scenario.append(scenario_roof + "+" + scenario_env)  # create the new total scenario
    return list_total_scenario

def make_all_technology_list(list_tech_roof, list_tech_env):
    """ Make the list of all technologies"""
    # Initialize the list with all the technologies
    list_tech = list_tech_roof
    for tech in list_tech_env:
        if tech not in list_tech:
            list_tech.append(tech)  # add the technology if not already in the initialized list of technology
    return list_tech

def calc_wb(nb_of_years, lifetime, shape):
    """ Calculate Weibull distribution values """
    list_wb = []
    for i in range(1, nb_of_years + 1):
        val = shape*i**(shape-1)/(lifetime**shape)*math.exp(-(i/lifetime)**shape)
        list_wb.append(val)
        
    return(list_wb)

def calc_wb_cumu(nb_of_years, lifetime, shape):
    """ calculate cumulative Weibull distribution values """
    list_wb_cumu = []
    for i in range(1, nb_of_years + 1):
        val = (1-np.exp(-(i / lifetime)**shape))
        list_wb_cumu.append(val)
        
    return(list_wb_cumu)
                                                             
def initialize_dic_result_per_year(nb_of_years, list_wb, list_level, list_b_type, list_tech):
    """ Make a dictionary with all the branches to collect the results for all the tested configuration """
    # Initialize the result dictionary
    dic_result_per_year = {}
    # loop over all the configuration parameters
    for year in range(nb_of_years):  # eventually +1 if we go up to year 100 (=101 years in total)
        dic_result_per_year[year] = {}  # Initialization (need say it's a dictionary)
        for b_type in list_b_type:
            dic_result_per_year[year][b_type] = {"roof_scenario": {}, "env_scenario": {},
                                                 "total_scenario": {}}  # Initialization
            # Roof
            dic_result_per_year[year][b_type]["roof_scenario"] = {}  # Initialization
            dic_result_per_year[year][b_type]["roof_scenario"][list_tech[0]] = {"stock": 0., "modules" : 0., "waste": 0., "waste_modules": 0., "waste_rounded": 0.}
            # Envelop
            dic_result_per_year[year][b_type]["env_scenario"] = {}  # Initialization
            dic_result_per_year[year][b_type]["env_scenario"][list_tech[1]] = {"stock": 0., "modules" : 0., "waste": 0., "waste_modules": 0., "waste_rounded": 0.}
            ## Total
            # make combinations of total scenarios
            dic_result_per_year[year][b_type]["total_scenario"] = {}  # Initialization
            for tech in list_tech:
                dic_result_per_year[year][b_type]["total_scenario"][tech] = {"stock": 0., "modules" : 0., "waste": 0., "waste_modules": 0., "waste_rounded": 0.}

    return dic_result_per_year

def initialize_dic_area(list_b_type):
    """ Make a dictionary to put in area data for stock calculation """
    # intialize area dictionary
    dic_area = {}
    for year in range(nb_of_years):
        dic_area[year] = {}
        for b_typ in list_b_type: 
            dic_area[year][b_typ] = {"roof_scenario": {}, "env_scenario": {}}
            
            # Roof
            dic_area[year][b_typ]["roof_scenario"] = {"area": 0.}
                
            # envelope
            dic_area[year][b_typ]["env_scenario"] = {"area" : 0.}
    
    return dic_area


def initialize_dic_energy(list_level, list_tech, list_b_type):
    """ Make a dictionary to put in energy production values """
    # intialize area dictionary
    dic_energy = {}
    for year in range(nb_of_years):
        dic_energy[year] = {}
        for b_type in list_b_type: 
            dic_energy[year][b_type] = {"roof_scenario": {}, "env_scenario": {}, "total_scenario" : {}}
            
            # Roof
            dic_energy[year][b_type]["roof_scenario"] = {}
            dic_energy[year][b_type]["roof_scenario"][list_tech[0]] = {"energy_production": 0.}
                
            # envelope
            dic_energy[year][b_type]["env_scenario"] = {}
            dic_energy[year][b_type]["env_scenario"][list_tech[1]] = {"energy_production": 0.}
                    
            ## Total
            # make list of all technologies
            dic_energy[year][b_type]["total_scenario"] = {}  # Initialization
            for tech in list_tech:
                dic_energy[year][b_type]["total_scenario"][tech] = {"energy_production": 0.}

    
    return dic_energy

def initialize_dic_irradiation(list_b_type):
    """ make a dictionary to put in irradiation values per scenario """
    dic_irradiation = {}
    
    for year in range(nb_of_years):
        dic_irradiation[year] = {}
        for b_type in list_b_type: 
            dic_irradiation[year][b_type] = {"roof_scenario": {}, "env_scenario": {}}
            
            # Roof
            dic_irradiation[year][b_type]["roof_scenario"] = {"irradiation": 0.}
                
            # envelope
            dic_irradiation[year][b_type]["env_scenario"] = {"irradiation" : 0.}
        
    return dic_irradiation

def calc_area_factor(dic_construction, construction_scenario, b_type):
    """ reducing area in early years of less existing construction """
    
    buildtype = b_type.split("_")[0] # only get name of b_type without angle of orientation
    if construction_scenario == "simple":
        dic_construction = dic_construction_simple
    elif construction_scenario == "phased":
        dic_construction = dic_construction_phased

    #parameters for construction processs
    total_nb_of_buildings = dic_construction[buildtype]["total_nb_of_buildings"]
    nb_buildings_parallel = dic_construction[buildtype]["nb_buildings_parallel"]
    construction_time = dic_construction[buildtype]["construction_time"]
    length = (total_nb_of_buildings-nb_buildings_parallel)/nb_buildings_parallel*construction_time

    if total_nb_of_buildings % nb_buildings_parallel != 0:
        phases = [i for i in range(0, round(length), construction_time)]
    elif total_nb_of_buildings % nb_buildings_parallel == 0:
        phases = [i for i in range(0, int(length) + 1, construction_time)]

    buildings = []
    nb_built = nb_buildings_parallel
    for year in range(1, nb_of_years + 1):
        if year in phases:
            if nb_built < total_nb_of_buildings <= nb_built + nb_buildings_parallel:
                nb_built = total_nb_of_buildings
            elif nb_built + nb_buildings_parallel < total_nb_of_buildings:
                nb_built = nb_built + nb_buildings_parallel
     
        elif year > phases[-1] + 1:
            nb_built = total_nb_of_buildings
            
        buildings.append(nb_built)
     
    buildings.insert(0, nb_buildings_parallel)
    buildings = buildings[0:100]
    area_factor = [i/total_nb_of_buildings for i in buildings]

    return area_factor
              
def calc_area(dic_area, construction_scenario, list_scenario, list_b_type, datasheet, rows, columns):
    """ creates array with all PV-area inputs for each scenario and writes the values in the dic_area """
    
    if construction_scenario == "simple":
        dic_construction = dic_construction_simple
    elif construction_scenario == "phased":
        dic_construction = dic_construction_phased

    arr_area = np.zeros((len(list_b_type), len(list_scenario)))
    for row in range(rows[0], rows[1]):
        for column in range(columns[0], columns[1]):
            if column == columns[0]:
                arr_area[row-rows[0],column-columns[0]] = datasheet.cell(row,column).value * area_factor_roof
            else:
                arr_area[row-rows[0],column-columns[0]] = datasheet.cell(row,column).value * area_factor_env
    r = 0
    c = 0
    for scenario in list_scenario:
        for b_type in list_b_type:
            # print(b_type)
            # print(scenario)
            factorlist = calc_area_factor(dic_construction, construction_scenario, b_type)
            # print(factorlist)
            for year in range(nb_of_years):
                if scenario == "roof_scenario":
                    dic_area[year][b_type]["roof_scenario"]["area"] = arr_area[r,c] * factorlist[year]
                else: 
                    dic_area[year][b_type]["env_scenario"]["area"] = arr_area[r, c] * factorlist[year]
            
            r = r + 1
        c = c + 1
        r = 0
    return arr_area

def calc_stocks(dic_area, level, dic_result_per_year, mass_conv):
    """ calculate stocks for each branch and write it in dic_results_per_year """
    # loop over all years
    for year in list(dic_result_per_year.keys()):
        for b_type in list(dic_result_per_year[year].keys()):
            for var in list(dic_result_per_year[year][b_type].keys()):       
                if var == "roof_scenario":
                    a = dic_area[year][b_type][var]["area"]
                    tech = list_tech[0]
                    stock = a*mass_conv 
                    dic_result_per_year[year][b_type][var][tech]["stock"] = stock 
                            
                elif var == "env_scenario":
                    a = dic_area[year][b_type][var]["area"]
                    stock = a*mass_conv 
                    dic_result_per_year[year][b_type][var][list_tech[1]]["stock"] = stock
                 
def calc_modules(dic_area, level, dic_result_per_year, area_to_module):
    
    for year in list(dic_result_per_year.keys()):
        for b_type in list(dic_result_per_year[year].keys()):
            for var in list(dic_result_per_year[year][b_type].keys()):
                for tech in list(dic_result_per_year[year][b_type][var].keys()):
                    if var == "total_scenario":
                        pass
                    else:
                        modules = math.ceil(dic_area[year][b_type][var]["area"] / area_to_module)
                        dic_result_per_year[year][b_type][var][tech]["modules"] = modules 
                        
def calc_cohorts(dic_result_per_year, b_type, var, tech, wb, level, nb_of_years):
    """ calculates an array of age-cohorts of outflows for each branch"""
    
    dic_wb = {"regular loss" : wb_cumu_RL, "early loss" : wb_cumu_EL}
    waste_arr = np.zeros((len(list_life), len(list_life)))
    waste_modules_arr = np.zeros((len(list_life), len(list_life)))
    list_inflow = []
    wb_dist = dic_wb[wb]
    
    # create waste arrays for each year
    for y in range(0, nb_of_years):  
        if y == 0:
            inflow_init = dic_result_per_year[0][b_type][var][tech]["stock"]
            modules_init = dic_result_per_year[0][b_type][var][tech]["modules"]
            list_inflow.append(float(inflow_init))
            module_list = []
            waste_list = []
            for i in range(nb_of_years):
                
                # in year 0 there is only inflow from construction, not from waste
                if i == 0:
                    waste = wb_dist[i]*inflow_init
                    waste_arr[y, i] = waste
                    waste_modules = math.ceil(waste /weight_to_module)
                    if waste > (weight_to_module * 1):
                        waste_modules_arr[y,i] = waste_modules
                elif i > 0:
                    waste = (wb_dist[i]-wb_dist[i-1])*inflow_init 
                    waste_arr[y, i] = waste
                    waste_modules = math.ceil(waste /weight_to_module)
                    module_list.append(waste_modules)
                    waste_list.append(waste)
                    if waste > (weight_to_module * 1):
                        waste_modules_arr[y,i] = waste_modules

            
        elif 1 <= y < 100:
            # inflow from replacement
            inflow_waste = np.sum(waste_arr, axis = 0)[y]
            
            # check for inflow from construction
            if dic_result_per_year[y][b_type][var][tech]["stock"] > dic_result_per_year[y-1][b_type][var][tech]["stock"]:
                inflow_construction = dic_result_per_year[y][b_type][var][tech]["stock"] - dic_result_per_year[y-1][b_type][var][tech]["stock"]
            else: 
                inflow_construction = 0
            
            inflow_total = inflow_waste + inflow_construction
            list_inflow.append(float(inflow_total))
            
            for i in list_life:
                x = i + y
                if i == 0:
                    waste = wb_dist[i]*inflow_total
                    waste_arr[y, x] = waste
                    waste_modules = math.ceil( waste / (weight_to_module))
                    if waste > (weight_to_module * 1):
                        waste_modules_arr[y,x] = waste_modules
                        
                elif i > 0 and x < 100:
                    waste = (wb_dist[i]-wb_dist[i-1])*inflow_total 
                    waste_arr[y, x] = waste
                    waste_modules = math.ceil(waste / weight_to_module)

                    if waste > (weight_to_module * 1):
                        waste_modules_arr[y,x] = waste_modules
            
    list_inflow_modules = [math.ceil(i/weight_to_module) for i in list_inflow]

    
    return waste_arr, waste_modules_arr, list_inflow, list_inflow_modules

def calc_left_modules(dic_result_per_year, year_test, b_type, var, tech, wb, level, nb_of_years):
    """calculates lists of emerging waste from each year for a specific test year"""
    
    waste_arr, waste_modules_arr, list_inflow, list_inflow_modules = calc_cohorts(dic_result_per_year, b_type, var, tech, wb, level, nb_of_years)
    left_weight_list = []
    left_module_list = []
    waste_list = []
    waste_module_list = []
    
    # go through each year and calculate the waste that is emerging from each preceding year
    for year in range(year_test + 1):
        if year == 0:
            inflow = dic_result_per_year[0][b_type][var][tech]["stock"]
            inflow_modules = dic_result_per_year[0][b_type][var][tech]["modules"]
        else:
            inflow_waste = np.sum(waste_arr, axis = 0)[year]
            inflow_modules_from_waste = np.sum(waste_modules_arr, axis = 0)[year]
            
            if dic_result_per_year[year][b_type][var][tech]["stock"] > dic_result_per_year[year-1][b_type][var][tech]["stock"]:
                inflow_construction = dic_result_per_year[year][b_type][var][tech]["stock"] - dic_result_per_year[year-1][b_type][var][tech]["stock"]
                inflow_modules_from_construction = dic_result_per_year[year][b_type][var][tech]["modules"] - dic_result_per_year[year-1][b_type][var][tech]["modules"]
                # print(inflow_modules_from_construction, year)
            else: 
                inflow_construction = 0
                inflow_modules_from_construction = 0
                
            inflow = inflow_waste + inflow_construction
            inflow_modules = inflow_modules_from_waste + inflow_modules_from_construction
        
        waste = 0
        waste_modules = 0
        # sum up all emerging waste from preceding years
        for i in range(year_test):
            waste += waste_arr[year, i]
            waste_modules += waste_modules_arr[year, i]

        left_from_year = inflow - waste
        left_weight_list.append(left_from_year)
        waste_list.append(waste)
        left_modules_from_year = inflow_modules - waste_modules
        left_module_list.append(left_modules_from_year)
        waste_module_list.append(waste_modules)
    
    return left_weight_list, left_module_list, waste_list, waste_module_list

def calc_cumu_waste(dic_result_per_year, b_type, var, tech, nb_of_years):
    """calculate the cumulative waste at the EOL of the building"""
    
    list_waste = []
    for i in range(nb_of_years + 1):
        waste = 0
        for x in range(i):
            waste += dic_result_per_year[x][b_type][var][tech]["waste"]
        list_waste.append(waste)
    
    return list_waste

def calc_waste(dic_result_per_year, wb, level):
    """ calculate waste for each branch and write it in dic_results_per_year"""

    # loop over all years
    for year in list(dic_result_per_year.keys()):
        for b_type in list(dic_result_per_year[year].keys()):
            for var in list(dic_result_per_year[year][b_type].keys()):
                if var == "roof_scenario":
                    tech = "rooftech"
                elif var == "env_scenario":
                    tech = "envtech"
                    
                if var == "total_scenario":
                    pass
                else:
                    arr, new_arr, list_inflow, list_inflow_modules = calc_cohorts(dic_result_per_year, b_type, var, tech, wb, level, nb_of_years) # create waste array to extract data from
                    w = np.sum(arr, axis = 0)[year] # waste is sum of all outflows from the previous age-cohorts in the considered year
                    if w > (weight_to_module * 1):
                        waste_modules = math.ceil(w / weight_to_module)
                    else:
                        waste_modules = 0
                    
                    if year == 99: # in the last year all modules become waste
                        
                        w = dic_result_per_year[year - 1][b_type][var][tech]["stock"]
                        waste_modules = math.ceil(w / weight_to_module)
                    
                    dic_result_per_year[year][b_type][var][tech]["waste"] = w # change value of waste in result dictionary
                    dic_result_per_year[year][b_type][var][tech]["waste_modules"] = waste_modules
                    dic_result_per_year[year][b_type][var][tech]["waste_rounded"] = waste_modules * weight_to_module

def compute_total_scenarios(dic_result_per_year):
    """ Compute the stock and the waste of combined roof and envelop scenarios (and eventually others if needed """
    
    #  loop over all he years ( we add list before to convert it as it's not exactly a list and we cannot do a loop on it
    for year in list(dic_result_per_year.keys()):
        for b_type in list(dic_result_per_year[year].keys()):
            for tech in list(dic_result_per_year[year][b_type]["total_scenario"].keys()):
                # Roof
                try:
                    dic_result_per_year[year][b_type]["roof_scenario"][tech]["stock"]
                except:
                    roof_stock = 0
                    roof_modules = 0
                    roof_waste = 0
                    roof_waste_modules = 0
                    roof_waste_rounded = 0
                else:
                    roof_stock = dic_result_per_year[year][b_type]["roof_scenario"][tech]["stock"]
                    roof_modules = dic_result_per_year[year][b_type]["roof_scenario"][tech]["modules"]
                    roof_waste = dic_result_per_year[year][b_type]["roof_scenario"][tech]["waste"]
                    roof_waste_modules = dic_result_per_year[year][b_type]["roof_scenario"][tech]["waste_modules"]
                    roof_waste_rounded = dic_result_per_year[year][b_type]["roof_scenario"][tech]["waste_rounded"]
                # Envelop
                try:
                    dic_result_per_year[year][b_type]["env_scenario"][tech]["stock"]
                except:
                    env_stock = 0
                    env_modules = 0
                    env_waste = 0
                    env_waste_modules = 0
                    env_waste_rounded = 0
                else:
                    env_stock = dic_result_per_year[year][b_type]["env_scenario"][tech]["stock"]
                    env_modules = dic_result_per_year[year][b_type]["env_scenario"][tech]["modules"]
                    env_waste = dic_result_per_year[year][b_type]["env_scenario"][tech]["waste"]
                    env_waste_modules = dic_result_per_year[year][b_type]["env_scenario"][tech]["waste_modules"]
                    env_waste_rounded = dic_result_per_year[year][b_type]["env_scenario"][tech]["waste_rounded"]
                # Total
                dic_result_per_year[year][b_type]["total_scenario"][tech][
                    "stock"] = roof_stock + env_stock
                dic_result_per_year[year][b_type]["total_scenario"][tech][
                    "modules"] = roof_modules + env_modules
                dic_result_per_year[year][b_type]["total_scenario"][tech][
                    "waste"] = roof_waste + env_waste
                dic_result_per_year[year][b_type]["total_scenario"][tech][
                    "waste_modules"] = roof_waste_modules + env_waste_modules
                dic_result_per_year[year][b_type]["total_scenario"][tech][
                    "waste_rounded"] = roof_waste_rounded + env_waste_rounded
                           
    return dic_result_per_year

def calc_waste_real(dic_result_per_year):
    """ panels will not get replaced yearly but after fixed intervals of 25 years """

    dic_real = copy.deepcopy(dic_result_per_year)
    list_waste_years = [25, 50, 75, 99]
    for year in list(dic_real.keys()):
        for b_type in list(dic_real[year].keys()):
            for var in list(dic_real[year][b_type].keys()):
                for tech in list(dic_real[year][b_type][var].keys()):
                    if year in list_waste_years:
                        dic_real[year][b_type][var][tech]["waste"] = dic_real[year][b_type][var][tech]["stock"]
                    else:
                        dic_real[year][b_type][var][tech]["waste"] = 0.
    return dic_real

def write_waste_data(dic_result_per_year, dic_real, real_or_ideal, weibull, level, nb_of_years):
    """writes cumulative waste at EOL from building in excel file"""
    
    ws = datafile.create_sheet("waste" + " " + weibull + " " + level)
    ws.cell(1,1).value = "Waste in tons"
    m = 2
    
    for tech in list_tech:
        n = 2
        for b_type in list_b_type:
            scenario = "total_scenario"
            if real_or_ideal == "ideal":
                reslist = calc_cumu_waste(dic_result_per_year, b_type, scenario, tech, nb_of_years)
                res = reslist[nb_of_years]
            elif real_or_ideal == "real":
                res = 0
                for year in range(nb_of_years):
                    res += dic_real[year][weibull][level][b_type]["total_scenario"][tech]["waste"]
            cell_name = ws.cell(n, 1)   
            cell_waste = ws.cell(n, m)
            cell_label = ws.cell(1, m)
            
            cell_name.value = b_type
            cell_waste.value = res
            cell_label.value = tech + " " + weibull[:3] + " " + level

            n = n+1
        m = m + 1
    
    datafile.save("dmfa model 2.xlsx")
    
def create_data(level, wb, construction):
    """runs all calculations and writes data in .json files"""
    
    dic_result_per_year = initialize_dic_result_per_year(nb_of_years, list_wb, list_level, list_b_type, list_tech)
    dic_area = initialize_dic_area(list_b_type)
    arr_area = calc_area(dic_area, construction, list_scenario, list_b_type, area_file, [4,5], [4,6])
    
    path = "area.json"
    with open(path, "w") as out_file:
        json.dump(dic_area, out_file, indent=4)

    path = "init.json"
    with open(path, "w") as out_file:
        json.dump(dic_result_per_year, out_file, indent=4)
        
    calc_stocks(dic_area, level, dic_result_per_year, mass_conv)
    calc_modules(dic_area, level, dic_result_per_year, area_to_module)
    
    path = "stocks.json"
    with open(path, "w") as out_file:
        json.dump(dic_result_per_year, out_file, indent=4)
        
    calc_waste(dic_result_per_year, wb, level)
    
    path = "waste.json"
    with open(path, "w") as out_file:
        json.dump(dic_result_per_year, out_file, indent=4)
        
    compute_total_scenarios(dic_result_per_year)
    
    path = "complete.json"
    with open(path, "w") as out_file:
        json.dump(dic_result_per_year, out_file, indent = 4)
        
    dic_real = calc_waste_real(dic_result_per_year)

    path = "compare.json"
    with open(path, "w") as out_file:
        json.dump(dic_result_per_year, out_file, indent = 4)

    path = "waste_real.json"
    with open(path, "w") as out_file:
        json.dump(dic_real, out_file, indent = 4)

    return dic_result_per_year, dic_real

""" PLOTS """

def plot_heatmap(dic_result_per_year, wb, b_type, level, var, tech, name, list_life, nb_of_years = nb_of_years):
    """ creates a plot to display the flow of age-cohorts thought time """
    
    waste_arr, new_arr, list_inflows, list_inflow_modules = calc_cohorts(dic_result_per_year, b_type, var, tech, wb, level, nb_of_years)
    
    fig, ax = plt.subplots()
    im = ax.imshow(waste_arr, cmap=plt.get_cmap('PuBuGn'), interpolation='nearest',
                    vmin=0, vmax=1)
    fig.colorbar(im)
    plt.title("Heat map" + name)
    ax.set_xlabel("Years")
    ax.set_ylabel("Age-cohorts")
    ax.set_xlim(0,nb_of_years)
    ax.set_ylim(nb_of_years,-1)
    plt.savefig('heatmap.png', dpi=1000)
    plt.show()

def plot_weibull(list_years, wb_pdf_RL, wb_pdf_EL, wb_cumu_RL, wb_cumu_EL):
    fig, ax = plt.subplots(figsize=(9, 6))
    plot1 = ax.plot(list_years, wb_pdf_RL, color = "k", label = "pdf_RL")
    plot2 = ax.plot(list_years, wb_pdf_EL, color = "k", linestyle = "dashed", label = "pdf_EL" )
    axy = ax.twinx()
    plot3 = axy.plot(list_years, wb_cumu_RL, color = "r", label = "cumu_RL")
    plot4 = axy.plot(list_years, wb_cumu_EL, color = "r", linestyle = "dashed", label = "cumu_EL")
    plt.title("Weibull distribution")
    ax.set_xlabel('Years')
    plt.rcParams.update({'font.size': 14})
    axy.set_ylabel('cumulative', color='r')
    ax.set_ylabel('probability density function', color='k')
    lns = plot1 + plot2 + plot3 + plot4
    labs = [l.get_label() for l in lns]
    ax.legend(lns, labs, loc="right")
    plt.show()

def plot_stocks(dic_result_per_year, b_type, var, scenario, tech, list_years):
    stocks = []
    ra = [i for i in range(0, nb_of_years)]
    for i in ra:
        stocks.append(dic_result_per_year[i]["regular loss"][b_type][var][scenario][tech]["stock"])
    
    fig2, ax2 = plt.subplots(figsize=(9, 6))
    ax2.plot(list_years[:], stocks, color = "k", label = str(tech))
    ax2.set_xlabel('Years')
    ax2.set_ylabel('Stocks (t)', color='k')
    plt.title("Stocks " + b_type + " " + " " + var + " " + tech)
    plt.rcParams.update({'font.size': 14})
    plt.legend(loc="upper left")
    plt.show

def plot_waste(dic_result_per_year, waste_or_modules, b_type, var, tech, nb_of_years):

    tech_waste = []
    ra = [i for i in range(nb_of_years)]
    for i in ra:
        tech_waste.append(dic_result_per_year[i][b_type][var][tech][waste_or_modules])
    # print(tech_waste)
    fig2, ax2 = plt.subplots(figsize=(9, 6))
    ax2.plot(ra, tech_waste, color = "k", label = str(tech))
    ax2.set_xlabel('Years')
    ax2.set_ylabel('Waste (t)', color='k')
    plt.title("Emerging waste " + b_type  + " " + tech)
    plt.rcParams.update({'font.size': 14})
    plt.legend(loc="upper left")
    plt.show
    
def plot_cumu_waste(dic_result_per_year, waste_or_modules, b_type, var, nb_of_years):
    
    if var == "roof_scenario":
        tech = "rooftech"
    else: 
        tech = "envtech"
    list_cumu = []
    xaxis = [i for i in range(nb_of_years + 1)]
    for i in range(nb_of_years + 1):
        waste = 0
        for x in range(i):
            waste += dic_result_per_year[x][b_type][var][tech][waste_or_modules]
        list_cumu.append(waste)
    print(list_cumu)
        
    fig2, ax2 = plt.subplots(figsize=(9, 6))
    ax2.plot(xaxis, list_cumu, color = "k", label = str(tech))
    ax2.set_xlabel('Years')
    if waste_or_modules == "waste":
        ax2.set_ylabel('Waste (t)', color='k')
    else:
        ax2.set_ylabel('Waste Modules', color='k')
    plt.title("Emerging cumulative waste " + b_type  + " " + tech)
    plt.rcParams.update({'font.size': 14})
    plt.legend(loc="upper left")
    plt.show
    

def plot_waste_for_wb(dic_result_per_year, b_type, var, scenario, tech, list_years, nb_of_years):

    all_data = []
    for x in list_wb:
        single_data = [dic_result_per_year[i][x][b_type][var][scenario][tech]["waste"] for i in range(0, nb_of_years)]
        all_data.append(single_data)
    
    single_data = calc_cumu_waste(dic_result_per_year, b_type, var, tech, nb_of_years)
    all_data.append(single_data)
    
    fig, ax = plt.subplots(figsize=(9, 6))
    plot1 = ax.plot(list_years, all_data[0], color = "k", label = "annual waste, regular loss")
    plot2 = ax.plot(list_years, all_data[1], color = "k", linestyle = "dashed", label = "annual waste, early loss" )
    axy = ax.twinx()
    plot3 = axy.plot(list_years, all_data[2], color = "cadetblue", label = "cumulative waste, regular loss")
    plot4 = axy.plot(list_years, all_data[3], color = "cadetblue", linestyle = "dashed", label = "cumulative waste, early loss")
    plt.title("Generated PV waste, annual and cumulative")
    ax.set_xlabel('Year')
    plt.rcParams.update({'font.size': 14})
    axy.set_ylabel('cumulative waste generation in tons', color='cadetblue')
    ax.set_ylabel('Waste generation in tons', color='k')
    lns = plot1 + plot2 + plot3 + plot4
    labs = [l.get_label() for l in lns]
    ax.legend(lns, labs, loc="lower right")
    plt.savefig('waste_wb.png', dpi=1000)
    plt.show()
    
def plot_total_waste(dic_result_per_year, b_type, scenario, list_years, list_tech, nb_of_years):

    all_data = []
    for tech in list_tech:
        single_data = calc_cumu_waste(dic_result_per_year, b_type, "total_scenario", tech, nb_of_years)
        all_data.append(single_data)
    
    fig, ax = plt.subplots(figsize=(9, 6))
    plot1 = ax.plot(list_years, all_data[0], color = "k", label = "c-Si waste, regular loss")
    plot2 = ax.plot(list_years, all_data[2], color = "k", linestyle = "dashed", label = "c-Si waste, early loss" )
    plot3 = ax.plot(list_years, all_data[1], color = "cadetblue", label = "thin film waste, regular loss")
    plot4 = ax.plot(list_years, all_data[3], color = "cadetblue", linestyle = "dashed", label = "thin film waste, early loss")
    plt.title("Generated total PV waste for both roof and facades")
    ax.set_xlabel('Year')
    plt.rcParams.update({'font.size': 14})
    ax.set_ylabel('Waste generation in tons', color='k')
    lns = plot1 + plot2 + plot3 + plot4
    labs = [l.get_label() for l in lns]
    ax.legend(lns, labs, loc="best")
    plt.savefig('waste.png', dpi=1000)
    plt.show()

def plot_waste_comp_real(dic_result_per_year, dic_real, year, var, tech):

    data_list = []
    buildinglist = list_b_type[0::4]
    dic_list = [dic_result_per_year, dic_real]
    for dic in dic_list:
        waste_list = []
        for b_type in buildinglist:
            waste_cumu = 0
            for i in range(year):
                waste_cumu = waste_cumu + dic[i][b_type][var][tech]["waste"]
            waste_list.append(waste_cumu)
        data_list.append(waste_list)

    y_axis = data_list
    x_axis = buildinglist
    label = [i[:-2] for i in x_axis]


    plt.rcParams["figure.figsize"] = [8, 6.50]
    plt.rcParams["figure.autolayout"] = True
    d = {"Building type": label, "ideal": y_axis[0], "real": y_axis[1]}
    my_colors = ["mediumaquamarine", "lightskyblue"]
    df = pd.DataFrame(d)
    df.set_index("Building type").plot(kind="bar", align="center", color=my_colors, legend=True,
                                       width=0.65)
    plt.tick_params(rotation=0)
    plt.title("Waste generation after " + str(year) + " years")
    plt.ylabel("Generated waste in tons")
    # plt.legend(loc = "best")
    plt.show()

def plot_energy(dic_energy, list_level, b_type, var, scenario, tech, list_years):
    all_data = []
    single_data = []
    for level in list_level:
        single_data = [dic_energy[i][level][b_type][var][scenario][tech]["energy_production"] for i in range(0, nb_of_years)]
        all_data.append(single_data)
    
    fig, ax = plt.subplots(figsize=(9, 6))
    plot1 = ax.plot(list_years, all_data[0], color = "k", label = "Energy Generation, efficient scenario")
    plot2 = ax.plot(list_years, all_data[1], color = "k", linestyle = "dashed", label = "Energy Generation, aesthetic scenario" )
    plt.title("Generated PV energy")
    ax.set_xlabel('Year')
    plt.rcParams.update({'font.size': 14})
    ax.set_ylabel('Energy in GWh', color='k')
    lns = plot1 + plot2
    labs = [l.get_label() for l in lns]
    ax.legend(lns, labs, loc="lower right")
    #plt.savefig('energyproduction.png', dpi=1000)
    plt.show()

def plot_energy_comp_btype(dic_energy, year, level, scenario):
    energylist = []
    blist = []
    x_axis = []
    y_axis = []
    for b_type in list_b_type:
        energy = 0
        for tech in list_tech:
            
            energy = energy + dic_energy[year][level][b_type]["total_scenario"][scenario][tech]["energy_production"]
        
        energylist.append(energy)
        blist.append(b_type)

    print(len(energylist))
    y_axis = [energylist[0::4], energylist[1::4], energylist[2::4], energylist[3::4], energylist[4::4]]
    x_axis = blist[0::4]
    label = [i[:-2] for i in x_axis]
    print(x_axis)
    print(label)
    plt.rcParams["figure.figsize"] = [8, 6.50]
    plt.rcParams["figure.autolayout"] = True
    d = {"Building type" : label, "0 ° rotation" : y_axis[0], "45 ° rotation" : y_axis[1], "90 ° rotation" : y_axis[2], "135 ° rotation" : y_axis[3]}
    my_colors = ["cadetblue", "lightskyblue", "slategray", "powderblue"]
    df = pd.DataFrame(d)
    df.set_index("Building type").plot(kind = "bar", align = "center", color = my_colors, legend = True, ylim = [0,18], width = 0.65)
    plt.tick_params(rotation = 0)
    plt.title("Energy generation per Year", fontweight = "bold", size=16)
    plt.ylabel("Generated energy \n in GWh")
    # plt.legend(loc = "best")
    plt.savefig('energy.png', dpi=1000)
    plt.show()
    
def plot_energy_comp_level(dic_energy, year, scenario):
    energylist = []
    buildinglist = list_b_type[0::4]
    x_axis = []
    y_axis = []
    for b_type in buildinglist:
        for level in list_level:
            energy = 0
            for tech in list_tech:
                energy = energy + dic_energy[year][level][b_type]["total_scenario"]["roof+envelop_800"][tech]["energy_production"]
            energylist.append(energy)
    y_axis = [energylist[0::2], energylist[1::2]]
    x_axis = buildinglist
    label = [i for i in x_axis]
    plt.rcParams["figure.figsize"] = [8, 6.50]
    plt.rcParams["figure.autolayout"] = True
    d = {"Building type" : label, "upper efficiency" : y_axis[0], "lower efficiency" : y_axis[1]}
    my_colors = ["mediumaquamarine", "lightskyblue"]
    df = pd.DataFrame(d)
    df.set_index("Building type").plot(kind = "bar", align = "center", color = my_colors, legend = True, ylim = [0,18], width = 0.65)
    plt.tick_params(rotation = 0)
    plt.title("Energy generation per year")
    plt.ylabel("Generated energy in GWh")
    # plt.legend(loc = "best")
    plt.savefig('energy_lev.png', dpi=1000)
    plt.show()

# che
# wb_cumu_RL = calc_wb_cumu(nb_of_years, lt_PV, wb_shape_RL)
# wb_cumu_EL = calc_wb_cumu(nb_of_years, lt_PV, wb_shape_EL)